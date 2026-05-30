"""Grid hyperparameter sweep for mrDMD-QR baseline variant on TPU.

Scans (L, max_cyc, r_max_per_dmd) over a multi-stage grid for the
multi-direction (11 AOAs) sensor placement task, and evaluates
Unknown_MAE at multiple sensor counts. Reuses load_data / build_basis /
reconstruct_l2 from the main script so any change in the upstream
pipeline propagates automatically.

Outputs (under mrdmd_qr/mode_result/sweep/<variant>_<snap>/):
  grid_summary.xlsx      one row per (L, max_cyc, r_max, n)
  grid_per_angle.xlsx    one row per (L, max_cyc, r_max, n, AOA)
  heatmap_n{N}.png       L x max_cyc heatmap of Unknown_MAE_Mean per r_max,
                         one figure per sensor count N

Note: mrDMD-QR is deterministic. Std reported here is over the 11 AOAs
(physical variability), not over repeated runs.
"""

import itertools
import time
from collections import defaultdict
from pathlib import Path
from types import SimpleNamespace

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.linalg import qr

# mrdmdqr_tpu_l2.py and mrdmd_utils.py both sit next to this script.
import sys
sys.path.insert(0, str(Path(__file__).resolve().parent))

from mrdmdqr_tpu_l2 import (
    MODE_RESULT_ROOT,
    RAW_DIR,
    TRAIN_RATIO,
    VALID_RATIO,
    _eval_global_sensors_no_centering,
    build_basis,
    load_data,
    reconstruct_l2,
)

# =====================================================================
# GRID — three-stage hyperparameter search history
# =====================================================================
# Stage 1 (coarse):
#     L_GRID  = [3, 5, 7]
#     MC_GRID = [3, 5, 7]
#     R_GRID  = [30, 50, 70, 100]   # R=70 added as a fill between 50 and 100
#     → 36 combos total (27 initial + 9 R=70 fill, merged into one table)
#     → best @ n=20: L=7, mc=5, R=30, Unknown MAE = 0.0987
#     → output: mrdmd_results/{02_nohankel_n2048, 03_nohankel_fullsnap_stage1}/
#       grid_summary.xlsx contains all 36 rows; the R=70 sweep below writes
#       to a temporary "_r70fill" folder then is merged in by the
#       post-processing script.
#
# Stage 2 (fine, 21 combos, 10 n values, ~45 min @ fullsnap):
#     L_GRID  = [6, 7, 8]
#     MC_GRID = [5]
#     R_GRID  = [10, 15, 20, 25, 30, 35, 40]
#       — initial Stage 2 (15): R ∈ {20, 25, 30, 35, 40}
#       — R=15 fill (3)
#       — R=10 fill (3)
#     SENSOR_COUNTS = [2, 4, 6, 8, 10, 12, 14, 16, 18, 20]  (10 n values)
#     → best @ n=20: L=8, R=10, Unknown MAE = 0.0848
#     → best @ mean: L=7, R=10, mean MAE = 0.1276
#     → output: mrdmd_results/04_nohankel_fullsnap_stage2/
#     → Tables: table_2.xlsx (per-variant), fullsnap_table_2.xlsx (cross)
#
# Stage 3 (micro-refinement around R=5, 9 combos, ~10 min @ fullsnap):
#     L_GRID  = [6, 7, 8]
#     MC_GRID = [5]
#     R_GRID  = [4, 5, 6]
#     SENSOR_COUNTS = [2, 4, 6, 8, 10, 12, 14, 16, 18, 20]
#     → 9 (L,C,R) combos x 10 n values = 90 runs
#     → best @ n=20: L=8, R=6, Unknown MAE = 0.0848
#     → best @ mean: L=7, R=5, mean MAE = 0.1216 (slightly better than Stage 2)
#     → output: mrdmd_results/05_nohankel_fullsnap_stage3/
#     → Tables: table_3.xlsx (per-variant), fullsnap_table_3.xlsx (cross)
# =====================================================================
# Active run: Stage 2 default (set N_SNAPS = 30000 for fullsnap or 2048)
L_GRID         = [6, 7, 8]
MC_GRID        = [5]
R_GRID         = [10, 15, 20, 25, 30, 35, 40]
SENSOR_COUNTS  = [2, 4, 6, 8, 10, 12, 14, 16, 18, 20]
STAGE_TAG_OVERRIDE = None         # auto: "" if 21+ combos else "_stage2"

# =====================================================================
# Variant selector — three mrDMD modes:
#   "textbook_no_hankel":  exact DMD + no orthonormalization + no Hankel
#                          (textbook minimal; paper baseline)
#   "fbdmd_ortho":         FB-DMD + orthonormalization + no Hankel (d=1)
#   "hankel":              FB-DMD + orthonormalization + Hankel d=30
# =====================================================================
VARIANT        = "textbook_no_hankel"
D_EMBED_HANKEL = 5    # Hankel time-delay depth (only used when VARIANT involves Hankel)

# Snapshot setting: cap is min(N_SNAPS, n_tr). 30000 > n_tr -> full-snap.
N_SNAPS        = 30000
AMP_QUANTILE   = 0.0

if VARIANT == "textbook_no_hankel":
    D_EMBED, NO_HANKEL, USE_FB, ORTHO = 30, True, False, False
    _v = "nohankel"
elif VARIANT == "fbdmd_ortho":
    D_EMBED, NO_HANKEL, USE_FB, ORTHO = 1, False, True, True
    _v = "fbdmd_ortho"
elif VARIANT == "hankel":   # FB-DMD + Hankel + L2
    D_EMBED, NO_HANKEL, USE_FB, ORTHO = D_EMBED_HANKEL, False, True, True
    _v = f"hankel_d{D_EMBED_HANKEL}"
elif VARIANT == "exact_hankel":   # Hankel + Exact DMD (no FB), no L2
    D_EMBED, NO_HANKEL, USE_FB, ORTHO = D_EMBED_HANKEL, False, False, False
    _v = f"exact_hankel_d{D_EMBED_HANKEL}"
else:
    raise ValueError(VARIANT)

_snap_tag = "fullsnap" if N_SNAPS >= 30000 else f"n{N_SNAPS}"
if STAGE_TAG_OVERRIDE is not None:
    _stage_tag = STAGE_TAG_OVERRIDE
elif len(L_GRID) * len(MC_GRID) * len(R_GRID) < 27:
    _stage_tag = "_stage2"
else:
    _stage_tag = ""
OUT_DIR = MODE_RESULT_ROOT / "sweep" / f"{_v}_{_snap_tag}{_stage_tag}"
# =====================================================================


def make_args(L, max_cyc, r_max):
    return SimpleNamespace(
        L=L, max_cyc=max_cyc, r_max_per_dmd=r_max,
        d_embed=D_EMBED, amp_quantile=AMP_QUANTILE,
        no_hankel=NO_HANKEL,
        use_fb=USE_FB,
        orthonormalize=ORTHO,
    )


def attach_extra_splits(data, n_snaps_train):
    """Extend load_data() output with (i) the validation slice
    flat[i, n_tr : n_tr + n_va] and (ii) train-set angle labels matching
    data['train_concat_raw'] layout. Mirrors the test_concat_raw / angle_te
    conventions so _eval_global_sensors_no_centering can be reused as-is."""
    grid = np.load(RAW_DIR / "cp_grid.npy")
    n_ang, T = grid.shape[:2]
    flat = grid.reshape(n_ang, T, -1).astype(np.float64)
    n_tr = int(round(T * TRAIN_RATIO))
    n_va = int(round(T * VALID_RATIO))
    pa_valid_raw, angle_va = [], []
    for i in range(n_ang):
        cp_va = flat[i, n_tr:n_tr + n_va]
        pa_valid_raw.append(cp_va.T.copy())
        angle_va.append(np.full(cp_va.shape[0], i, dtype=np.int32))
    data["pa_valid_raw"]     = pa_valid_raw
    data["valid_concat_raw"] = np.concatenate(pa_valid_raw, axis=1)
    data["angle_va"]         = np.concatenate(angle_va)

    angle_tr = [np.full(X.shape[1], i, dtype=np.int32)
                for i, X in enumerate(data["pa_train_raw"])]
    data["angle_tr"]         = np.concatenate(angle_tr)
    return data


def run_one_combo(data, L, max_cyc, r_max):
    """Build basis + QR + evaluate at all sensor counts. Returns
    (summary_rows, per_angle_rows). Empty if basis build fails."""
    args = make_args(L, max_cyc, r_max)
    label = f"L{L}_mc{max_cyc}_r{r_max}"
    train_list = [data["train_concat_raw"]]

    t0 = time.perf_counter()
    try:
        Phi, info = build_basis(args, train_list, label, data["n_taps"])
    except Exception as e:
        print(f"  [{label}] FAILED build_basis: {e}")
        return [], []
    build_time = time.perf_counter() - t0
    r = info["basis_rank"]
    if r == 0:
        print(f"  [{label}] empty basis, skipping")
        return [], []

    _, _, piv = qr(Phi.T, pivoting=True, mode="economic")
    angles = sorted(np.unique(data["angle_va"]).tolist())

    summary_rows, per_angle_rows = [], []
    for n in SENSOR_COUNTS:
        if n > Phi.shape[0]:
            continue
        sel = np.sort(piv[:n])

        t1 = time.perf_counter()
        per_t, per_u, _ = _eval_global_sensors_no_centering(
            sel, data["n_taps"], Phi,
            data["valid_concat_raw"], data["angle_va"],
            reconstruct_l2, {})
        recon_time = time.perf_counter() - t1

        t_arr = np.array(per_t); u_arr = np.array(per_u)
        summary_rows.append({
            "L": L, "max_cyc": max_cyc, "r_max_per_dmd": r_max,
            "Sensor_Count": int(n), "Basis_Rank": int(r),
            "Valid_Total_MAE_Mean":   float(t_arr.mean()),
            "Valid_Total_MAE_Std":    float(t_arr.std()),
            "Valid_Unknown_MAE_Mean": float(u_arr.mean()),
            "Valid_Unknown_MAE_Std":  float(u_arr.std()),
            "Valid_Worst_AOA_Unknown_MAE": float(u_arr.max()),
            "Build_Time_s": float(build_time),
            "Recon_Time_s": float(recon_time),
        })
        for a, t_a, u_a in zip(angles, per_t, per_u):
            per_angle_rows.append({
                "L": L, "max_cyc": max_cyc, "r_max_per_dmd": r_max,
                "Sensor_Count": int(n), "Angle_ID": int(a),
                "Valid_Total_MAE": float(t_a),
                "Valid_Unknown_MAE": float(u_a),
            })
        print(f"  n={n:3d}  r={r:3d}  validUnkMAE={u_arr.mean():.6f}  "
              f"worst={u_arr.max():.6f}  t={recon_time:5.2f}s")
    return summary_rows, per_angle_rows


def evaluate_best_on_test(df, data):
    """For each sensor count, find the (L, max_cyc, r_max) combo with the
    lowest Valid_Unknown_MAE_Mean, then evaluate that selection on the
    held-out test set. Basis is rebuilt once per unique combo to avoid
    caching all 27 bases in memory."""
    sensor_counts = sorted(df["Sensor_Count"].unique())
    winners = []
    for n in sensor_counts:
        sub = df[df["Sensor_Count"] == n]
        best = sub.loc[sub["Valid_Unknown_MAE_Mean"].idxmin()]
        winners.append((int(n), int(best["L"]), int(best["max_cyc"]),
                        int(best["r_max_per_dmd"]),
                        float(best["Valid_Unknown_MAE_Mean"])))

    by_combo = defaultdict(list)
    for n, L, mc, r_max, vmae in winners:
        by_combo[(L, mc, r_max)].append((n, vmae))

    rows = []
    for (L, mc, r_max), n_list in by_combo.items():
        args = make_args(L, mc, r_max)
        Phi, info = build_basis(args, [data["train_concat_raw"]],
                                f"TEST_BEST_L{L}_mc{mc}_r{r_max}",
                                data["n_taps"])
        r = info["basis_rank"]
        _, _, piv = qr(Phi.T, pivoting=True, mode="economic")
        for n, valid_mae in n_list:
            if n > Phi.shape[0]:
                continue
            sel = np.sort(piv[:n])
            per_t, per_u, _ = _eval_global_sensors_no_centering(
                sel, data["n_taps"], Phi,
                data["test_concat_raw"], data["angle_te"],
                reconstruct_l2, {})
            t_arr = np.array(per_t); u_arr = np.array(per_u)
            rows.append({
                "Sensor_Count": int(n),
                "Best_L": L, "Best_max_cyc": mc,
                "Best_r_max_per_dmd": r_max, "Basis_Rank": int(r),
                "Valid_Unknown_MAE_Mean": float(valid_mae),
                "Test_Total_MAE_Mean":   float(t_arr.mean()),
                "Test_Total_MAE_Std":    float(t_arr.std()),
                "Test_Unknown_MAE_Mean": float(u_arr.mean()),
                "Test_Unknown_MAE_Std":  float(u_arr.std()),
                "Test_Worst_AOA_Unknown_MAE": float(u_arr.max()),
            })
            print(f"  [best@n={n}] L={L} mc={mc} r={r_max}  "
                  f"valid={valid_mae:.6f}  "
                  f"test={u_arr.mean():.6f}  worst={u_arr.max():.6f}")
    rows.sort(key=lambda x: x["Sensor_Count"])
    return pd.DataFrame(rows)


def plot_heatmaps(df, out_dir):
    sensor_counts = sorted(df["Sensor_Count"].unique())
    r_values      = sorted(df["r_max_per_dmd"].unique())
    L_values      = sorted(df["L"].unique())
    mc_values     = sorted(df["max_cyc"].unique())

    for n in sensor_counts:
        fig, axes = plt.subplots(1, len(r_values),
                                 figsize=(4.4 * len(r_values), 4.4))
        if len(r_values) == 1:
            axes = [axes]
        vmin = df[df["Sensor_Count"] == n]["Valid_Unknown_MAE_Mean"].min()
        vmax = df[df["Sensor_Count"] == n]["Valid_Unknown_MAE_Mean"].max()
        for ax, r_max in zip(axes, r_values):
            sub = df[(df["Sensor_Count"] == n) &
                     (df["r_max_per_dmd"] == r_max)]
            pivot = sub.pivot(index="L", columns="max_cyc",
                              values="Valid_Unknown_MAE_Mean")
            pivot = pivot.reindex(index=L_values, columns=mc_values)
            im = ax.imshow(pivot.values, cmap="viridis_r",
                           aspect="auto", vmin=vmin, vmax=vmax)
            ax.set_xticks(range(len(mc_values)))
            ax.set_xticklabels(mc_values)
            ax.set_yticks(range(len(L_values)))
            ax.set_yticklabels(L_values)
            ax.set_xlabel("max_cyc")
            ax.set_ylabel("L")
            ax.set_title(f"r_max_per_dmd = {r_max}")
            for i, L in enumerate(L_values):
                for j, mc in enumerate(mc_values):
                    val = pivot.iat[i, j]
                    if np.isfinite(val):
                        ax.text(j, i, f"{val:.4f}", ha="center",
                                va="center", color="white", fontsize=9)
            fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
        fig.suptitle(f"mrDMD-QR baseline grid sweep "
                     f"(Unknown MAE, n={n} sensors)", fontsize=12)
        fig.tight_layout()
        fig.savefig(out_dir / f"heatmap_n{n}.png", dpi=300)
        plt.close(fig)


def plot_n_vs_mae(df, out_dir):
    """One curve per (L, max_cyc, r_max). Useful overview of convergence."""
    fig, ax = plt.subplots(figsize=(9, 6))
    for (L, mc, r_max), grp in df.groupby(["L", "max_cyc", "r_max_per_dmd"]):
        grp = grp.sort_values("Sensor_Count")
        ax.plot(grp["Sensor_Count"], grp["Valid_Unknown_MAE_Mean"],
                marker="o", lw=0.8, ms=4,
                label=f"L={L}, mc={mc}, r={r_max}")
    ax.set_xlabel("Number of sensors n")
    ax.set_ylabel("Valid Unknown MAE (mean over 11 AOAs)")
    ax.set_title("mrDMD-QR baseline: sensor count vs validation error")
    ax.legend(fontsize=7, ncol=3, loc="upper right")
    ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_dir / "n_vs_mae_all_combos.png", dpi=300)
    plt.close(fig)


def write_gz_style(df, out_path, ns_label):
    """Reformat grid_summary into Guangzhou table:
        Config | L | R | C | ns | Basis_Rank | n=4 | n=8 | ... | n=20
    Bold + green-highlight the minimum value in each sensor-count column.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if df.empty or "Sensor_Count" not in df.columns:
        print(f"[skip] gz-style: grid_summary empty (no successful combos)")
        return

    sensor_counts = sorted(df["Sensor_Count"].unique())
    rows = []
    for (L, mc, r_max), grp in df.groupby(["L", "max_cyc", "r_max_per_dmd"]):
        grp = grp.sort_values("Sensor_Count")
        row = {
            "Config":     f"L={L} R={r_max} C={mc} ns={ns_label}",
            "L":          int(L),
            "R":          int(r_max),
            "C":          int(mc),
            "ns":         ns_label,
            "Basis_Rank": int(grp["Basis_Rank"].iloc[0]),
        }
        for n in sensor_counts:
            sub = grp[grp["Sensor_Count"] == n]
            row[f"n={n}"] = float(sub["Valid_Unknown_MAE_Mean"].iloc[0]) \
                if len(sub) else float("nan")
        rows.append(row)
    df_out = pd.DataFrame(rows).sort_values(["L", "R", "C"]).reset_index(drop=True)
    df_out.to_excel(out_path, index=False, sheet_name="grid")

    wb = load_workbook(out_path)
    ws = wb["grid"]
    bold      = Font(bold=True)
    header    = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="305496")
    best_fill = PatternFill("solid", fgColor="C6EFCE")
    border    = Border(left=Side(style="thin", color="BFBFBF"),
                       right=Side(style="thin", color="BFBFBF"),
                       top=Side(style="thin", color="BFBFBF"),
                       bottom=Side(style="thin", color="BFBFBF"))
    align_center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font, cell.fill, cell.alignment, cell.border = header, hdr_fill, align_center, border
    n_cols_start = 7
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment, cell.border = align_center, border
            if cell.column >= n_cols_start:
                cell.number_format = "0.0000"
    for n in sensor_counts:
        col_name = f"n={n}"
        col_idx = list(df_out.columns).index(col_name) + 1
        vals = df_out[col_name].values
        if all(pd.isna(vals)):
            continue
        min_val = min(v for v in vals if not pd.isna(v))
        for i, v in enumerate(vals, start=2):
            if not pd.isna(v) and abs(v - min_val) < 1e-12:
                c = ws.cell(row=i, column=col_idx)
                c.font, c.fill = bold, best_fill
    widths = {1: 22, 2: 4, 3: 6, 4: 4, 5: 10, 6: 12}
    for col_idx in range(7, 7 + len(sensor_counts)):
        widths[col_idx] = 10
    for col_idx, w in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    wb.save(out_path)
    print(f"[save] gz-style -> {out_path}")


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    combos = list(itertools.product(L_GRID, MC_GRID, R_GRID))
    n_combos = len(combos)
    print(f"[grid] {n_combos} combinations x {len(SENSOR_COUNTS)} "
          f"sensor counts = {n_combos * len(SENSOR_COUNTS)} evaluations")
    print(f"[grid] L={L_GRID}, max_cyc={MC_GRID}, r_max={R_GRID}")
    print(f"[grid] sensor_counts={SENSOR_COUNTS}")
    print(f"[grid] out_dir={OUT_DIR}\n")

    print("[load] loading TPU data ...")
    data = load_data(N_SNAPS)
    data = attach_extra_splits(data, N_SNAPS)
    print(f"  n_angles={data['n_ang']}, n_taps={data['n_taps']}")
    print(f"  train (mrDMD input): {data['train_concat_raw'].shape[1]} cols "
          f"(= {N_SNAPS} per angle x {data['n_ang']} angles)")
    print(f"  valid (sweep eval):  {data['valid_concat_raw'].shape[1]} cols")
    print(f"  test  (best-only):   {data['test_concat_raw'].shape[1]} cols\n")

    all_summary, all_per_angle = [], []
    t_global = time.perf_counter()
    for i, (L, mc, r_max) in enumerate(combos):
        print(f"[{i+1}/{n_combos}] L={L} max_cyc={mc} r_max={r_max}", flush=True)
        s_rows, pa_rows = run_one_combo(data, L, mc, r_max)
        all_summary.extend(s_rows)
        all_per_angle.extend(pa_rows)
        # Incremental save after every combo so partial runs are usable.
        pd.DataFrame(all_summary).to_excel(
            OUT_DIR / "grid_summary_partial.xlsx", index=False)
        pd.DataFrame(all_per_angle).to_excel(
            OUT_DIR / "grid_per_angle_partial.xlsx", index=False)
        elapsed = time.perf_counter() - t_global
        eta = elapsed / (i + 1) * (n_combos - i - 1)
        print(f"   [elapsed {elapsed/60:.1f} min, ETA {eta/60:.1f} min] "
              f"[saved partial {len(all_summary)} rows]\n", flush=True)

    df = pd.DataFrame(all_summary)
    df_pa = pd.DataFrame(all_per_angle)
    df.to_excel(OUT_DIR / "grid_summary.xlsx", index=False)
    df_pa.to_excel(OUT_DIR / "grid_per_angle.xlsx", index=False)
    df.to_csv(OUT_DIR / "grid_summary.csv", index=False)
    print(f"[save] summary -> {OUT_DIR / 'grid_summary.xlsx'}")
    print(f"[save] per-angle -> {OUT_DIR / 'grid_per_angle.xlsx'}")

    ns_label = min(N_SNAPS, 26214)
    write_gz_style(df, OUT_DIR / "table.xlsx", ns_label=ns_label)

    print("\n[test] evaluating best-valid combo per sensor count on test set ...")
    df_test = evaluate_best_on_test(df, data)
    df_test.to_excel(OUT_DIR / "best_on_test.xlsx", index=False)
    print(f"[save] best-on-test -> {OUT_DIR / 'best_on_test.xlsx'}")

    print("[plot] heatmaps + curves ...")
    plot_heatmaps(df, OUT_DIR)
    plot_n_vs_mae(df, OUT_DIR)

    # Sweep finished successfully — partial files are redundant with final.
    for fname in ("grid_summary_partial.xlsx", "grid_per_angle_partial.xlsx"):
        p = OUT_DIR / fname
        if p.exists():
            p.unlink()
            print(f"[clean] removed {p.name}")

    print(f"[done] total {(time.perf_counter() - t_global)/60:.1f} min")


if __name__ == "__main__":
    main()
